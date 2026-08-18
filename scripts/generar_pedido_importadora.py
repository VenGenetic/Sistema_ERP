"""
generar_pedido_importadora.py

Genera un pedido para la importadora en Excel, cruzando dos fuentes:

  1. odoo_scraper/data_costos_cantidad.json — el catálogo completo de la
     importadora (código, nombre, marca, costo con y sin IVA, cantidad
     disponible). Es el mismo archivo que lee scripts/sync-importer-stock.js
     para actualizar products.importer_stock, así que refleja el scrapeo más
     reciente. (public/pedidos.json, que usa la pantalla de Abastecimiento del
     sistema, quedó desactualizado — mayo de 2026 — y por eso no se usa aquí.)

  2. product_demands en Supabase — lo que los clientes han pedido y sigue
     activo (pending_stock o stock_available; se excluyen notificados,
     cancelados y vencidos, que ya no son una espera real).

Cruza ambas por código de repuesto para que la hoja de la importadora muestre
de un vistazo cuáles tienen clientes esperando, y los sube al principio de la
lista: son los que hay que priorizar al armar el pedido.

Uso:
    py scripts/generar_pedido_importadora.py
    py scripts/generar_pedido_importadora.py --out mi_pedido.xlsx

Requiere: pip install requests openpyxl
"""

import argparse
import json
import os
import sys
from collections import defaultdict
from datetime import date, datetime

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
IMPORTER_JSON = os.path.normpath(os.path.join(ROOT, "..", "odoo_scraper", "data_costos_cantidad.json"))

# Igual que en ExportDemandsModal.tsx: una demanda sigue "activa" mientras
# espera stock o el stock ya llegó y falta notificar al cliente. Notificado,
# cancelado y vencido ya no son una espera real.
ACTIVE_STATUSES = ("pending_stock", "stock_available")
STATUS_LABEL = {
    "pending_stock": "Esperando stock",
    "stock_available": "Stock disponible",
}

AMBER = "F59E0B"
INK = "0F172A"
PRIORITY_FILL = PatternFill("solid", fgColor="FEF3C7")
HEADER_FILL = PatternFill("solid", fgColor=AMBER)
HEADER_FONT = Font(bold=True, color=INK)
NOTE_FONT = Font(italic=True, size=9, color="6B7280")
MONEY_FMT = '"$"#,##0.00'
INT_FMT = "#,##0"


def load_env(path):
    values = {}
    if not os.path.exists(path):
        return values
    with open(path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            key, raw = line.split("=", 1)
            values[key.strip()] = raw.strip().strip("\"'")
    return values


ENV = load_env(os.path.join(ROOT, ".env"))
SUPABASE_URL = ENV.get("VITE_SUPABASE_URL")
SERVICE_KEY = ENV.get("SUPABASE_SERVICE_ROLE_KEY") or ENV.get("VITE_SUPABASE_SERVICE_ROLE_KEY")

if not SUPABASE_URL or not SERVICE_KEY:
    sys.exit("Faltan VITE_SUPABASE_URL o SUPABASE_SERVICE_ROLE_KEY en .env")

AUTH_HEADERS = {"apikey": SERVICE_KEY, "Authorization": f"Bearer {SERVICE_KEY}"}


def norm_sku(sku):
    return (sku or "").strip().upper()


# ─── Fuente 1: catálogo de la importadora ───

def load_importer_catalog():
    if not os.path.exists(IMPORTER_JSON):
        sys.exit(
            f"No se encontró {IMPORTER_JSON}.\n"
            "Es el archivo que arma el scraper de Odoo y que usa "
            "scripts/sync-importer-stock.js; sin él no hay de dónde sacar el "
            "stock de la importadora."
        )

    mtime = datetime.fromtimestamp(os.path.getmtime(IMPORTER_JSON))
    with open(IMPORTER_JSON, "r", encoding="utf-8") as f:
        raw = json.load(f)

    items = {}
    for item in raw:
        sku = norm_sku(item.get("codigo_referencia") or item.get("id"))
        if not sku:
            continue
        qty = int(item.get("stock_cantidad") or 0)
        # Se guarda el catálogo completo (para cruzar contra las demandas),
        # aunque a la hoja de la importadora sólo pase lo que tiene stock > 0.
        items[sku] = {
            "sku": sku,
            "name": item.get("nombre") or "",
            "brand": item.get("marca") or "",
            "category": item.get("categoria") or "",
            "qty": qty,
            "cost_no_vat": float(item.get("costo_sin_iva") or 0),
            "cost_vat": float(item.get("costo_con_iva") or 0),
        }
    return items, mtime


# ─── Fuente 2: solicitudes de clientes ───

def fetch_active_demands():
    statuses = ",".join(ACTIVE_STATUSES)
    select = (
        "id,phone_number,customer_name,status,created_at,"
        "product:products(id,sku,name,price,cost_without_vat,vat_percentage,"
        "importer_stock,inventory_levels(current_stock))"
    )
    url = (
        f"{SUPABASE_URL}/rest/v1/product_demands"
        f"?select={select}&status=in.({statuses})&order=created_at.asc"
    )
    res = requests.get(url, headers=AUTH_HEADERS, timeout=60)
    if not res.ok:
        sys.exit(f"No se pudieron leer las solicitudes de clientes: {res.status_code} {res.text}")
    return [d for d in res.json() if d.get("product")]


def r2(x):
    return round((x or 0) + 1e-9, 2)


# ─── Excel ───

def style_header(ws, ncols, note=None):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}1"
    if note:
        note_cell = ws.cell(row=1, column=ncols + 2, value=note)
        note_cell.font = NOTE_FONT


def autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_summary_sheet(wb, importer_mtime, n_importer, n_products_demand, n_demands):
    ws = wb.active
    ws.title = "Resumen"
    rows = [
        ("Pedido a la importadora", ""),
        ("Generado el", datetime.now().strftime("%d/%m/%Y %H:%M")),
        ("", ""),
        ("Artículos con stock en la importadora", n_importer),
        ("Productos con clientes esperando", n_products_demand),
        ("Solicitudes activas de clientes", n_demands),
        ("", ""),
        ("Catálogo de la importadora actualizado el", importer_mtime.strftime("%d/%m/%Y %H:%M")),
    ]
    for r, (label, value) in enumerate(rows, start=1):
        ws.cell(row=r, column=1, value=label).font = Font(bold=True) if label else Font()
        ws.cell(row=r, column=2, value=value)
    autosize(ws, [42, 24])

    stale_days = (datetime.now() - importer_mtime).days
    if stale_days > 3:
        warn = ws.cell(
            row=10, column=1,
            value=(
                f"Aviso: el catálogo de la importadora tiene {stale_days} días. "
                "Corre el scraper de Odoo antes de pedir si necesitas cifras del día."
            ),
        )
        warn.font = Font(color="B45309", italic=True)


def build_importer_sheet(wb, importer_items, waiting_by_sku):
    ws = wb.create_sheet("Stock Importadora")
    headers = ["Código", "Marca", "Nombre", "Categoría", "Cantidad Disponible",
               "Costo Unit. s/IVA ($)", "Costo Unit. c/IVA ($)", "Clientes Esperando"]
    ws.append(headers)

    rows = [
        (it["sku"], it["brand"], it["name"], it["category"], it["qty"],
         it["cost_no_vat"], it["cost_vat"], len(waiting_by_sku.get(it["sku"], [])))
        for it in importer_items.values()
        if it["qty"] > 0
    ]
    # Prioridad: primero lo que tiene clientes esperando (más clientes primero);
    # dentro de cada grupo, alfabético.
    rows.sort(key=lambda r: (-r[7], r[2]))

    for row in rows:
        ws.append(list(row))
        r = ws.max_row
        ws.cell(row=r, column=5).number_format = INT_FMT
        ws.cell(row=r, column=6).number_format = MONEY_FMT
        ws.cell(row=r, column=7).number_format = MONEY_FMT
        if row[7] > 0:
            for c in range(1, len(headers) + 1):
                ws.cell(row=r, column=c).fill = PRIORITY_FILL

    style_header(ws, len(headers), note="Resaltado = tiene clientes esperando")
    autosize(ws, [15, 16, 55, 18, 18, 18, 18, 16])
    return rows


def build_demand_summary_sheet(wb, waiting_by_sku, importer_items):
    ws = wb.create_sheet("Pedidos de Clientes")
    headers = ["Código", "Nombre", "Clientes Esperando", "Stock Local",
               "Stock Importadora", "Disponible en Importadora",
               "Costo Importadora s/IVA ($)", "Costo Importadora c/IVA ($)", "PVP Nuestro ($)"]
    ws.append(headers)

    rows = []
    for sku, items in waiting_by_sku.items():
        prod = items[0]["product"]
        local_stock = sum((lvl.get("current_stock") or 0) for lvl in (prod.get("inventory_levels") or []))
        imp = importer_items.get(sku)
        importer_qty = imp["qty"] if imp else int(prod.get("importer_stock") or 0)
        cost_no_vat = imp["cost_no_vat"] if imp else (prod.get("cost_without_vat") or 0)
        cost_vat = imp["cost_vat"] if imp else r2(cost_no_vat * (1 + (prod.get("vat_percentage") or 15.0) / 100))
        pvp = prod.get("price") or 0
        rows.append((
            sku, prod.get("name") or "", len(items), local_stock,
            importer_qty, "Sí" if importer_qty > 0 else "No",
            r2(cost_no_vat), r2(cost_vat), r2(pvp),
        ))

    rows.sort(key=lambda r: -r[2])
    for row in rows:
        ws.append(list(row))
        r = ws.max_row
        for col in (4, 5):
            ws.cell(row=r, column=col).number_format = INT_FMT
        for col in (7, 8, 9):
            ws.cell(row=r, column=col).number_format = MONEY_FMT
        if row[5] == "No":
            for c in range(1, len(headers) + 1):
                ws.cell(row=r, column=c).font = Font(color="9CA3AF")

    style_header(ws, len(headers), note='"No" en gris = no hay ahora en la importadora')
    autosize(ws, [15, 55, 18, 12, 16, 20, 22, 22, 14])
    return rows


def build_demand_detail_sheet(wb, demands):
    ws = wb.create_sheet("Detalle de Clientes")
    headers = ["Código", "Nombre Producto", "Cliente", "Teléfono", "Estado", "Fecha de Solicitud"]
    ws.append(headers)

    rows = []
    for d in demands:
        prod = d["product"]
        rows.append((
            prod.get("sku") or "",
            prod.get("name") or "",
            d.get("customer_name") or "Sin nombre",
            d.get("phone_number") or "",
            STATUS_LABEL.get(d.get("status"), d.get("status")),
            (d.get("created_at") or "")[:10],
        ))
    rows.sort(key=lambda r: (r[0], r[5]))
    for row in rows:
        ws.append(list(row))

    style_header(ws, len(headers))
    autosize(ws, [15, 50, 24, 16, 18, 16])


def main():
    parser = argparse.ArgumentParser(description=__doc__.split("\n")[1])
    parser.add_argument("--out", help="Ruta del archivo de salida")
    args = parser.parse_args()

    print("Cargando catálogo de la importadora...")
    importer_items, importer_mtime = load_importer_catalog()
    stale_days = (datetime.now() - importer_mtime).days
    print(f"  {len(importer_items)} artículos (actualizado hace {stale_days} día{'s' if stale_days != 1 else ''})")
    if stale_days > 3:
        print(f"  Aviso: el archivo tiene {stale_days} días. Considera correr el scraper de Odoo antes de pedir.")

    print("Leyendo solicitudes activas de clientes...")
    demands = fetch_active_demands()
    print(f"  {len(demands)} solicitudes activas ({', '.join(ACTIVE_STATUSES)})")

    waiting_by_sku = defaultdict(list)
    for d in demands:
        sku = norm_sku(d["product"].get("sku"))
        if sku:
            waiting_by_sku[sku].append(d)

    wb = Workbook()
    importer_rows = build_importer_sheet(wb, importer_items, waiting_by_sku)
    demand_rows = build_demand_summary_sheet(wb, waiting_by_sku, importer_items)
    build_demand_detail_sheet(wb, demands)
    build_summary_sheet(wb, importer_mtime, len(importer_rows), len(demand_rows), len(demands))
    wb.move_sheet("Resumen", offset=-3)  # la deja primera sin recrearla

    date_str = date.today().isoformat()
    out_path = args.out or os.path.join(ROOT, f"pedido_importadora_{date_str}.xlsx")
    wb.save(out_path)

    print(f"\nListo: {out_path}")
    print(f"  Stock Importadora:   {len(importer_rows)} artículos con stock")
    print(f"  Pedidos de Clientes: {len(demand_rows)} productos distintos")
    print(f"  Detalle de Clientes: {len(demands)} solicitudes")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        sys.exit("\nInterrumpido.")
